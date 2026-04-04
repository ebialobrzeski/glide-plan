"""Import service — file upload, preview, and confirm for external flight data."""
from __future__ import annotations

import csv
import hashlib
import io
import logging
import re
import uuid
from datetime import date, datetime, time, timezone
from typing import Optional

from sqlalchemy.dialects.postgresql import insert
from sqlalchemy.orm import Session

from backend.models.flight import Flight
from backend.models.import_log import ImportLog
from backend.models.user import User

logger = logging.getLogger(__name__)

SUPPORTED_SOURCES = ('leonardo', 'seeyou', 'weglide', 'excel', 'igc', 'csv')


class ImportServiceError(Exception):
    pass


# ── Parsers ──────────────────────────────────────────────────────────────────

def _parse_csv_generic(content: bytes) -> list[dict]:
    """Attempt to parse a generic CSV with auto-detected delimiter."""
    text = content.decode('utf-8-sig', errors='replace')
    dialect = csv.Sniffer().sniff(text[:2048], delimiters=',;\t')
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    return [dict(row) for row in reader]


def _parse_igc(content: bytes) -> list[dict]:
    """Parse an IGC file and return a single-flight dict."""
    text = content.decode('ascii', errors='replace')
    lines = text.splitlines()

    flight_date: Optional[date] = None
    aircraft_type = ''
    aircraft_reg = ''
    b_records: list[str] = []

    for line in lines:
        if line.startswith('HFDTE'):
            m = re.search(r'HFDTE(?:DATE:)?(\d{2})(\d{2})(\d{2})', line)
            if m:
                dd, mm, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
                year = 2000 + yy if yy < 70 else 1900 + yy
                flight_date = date(year, mm, dd)
        elif line.startswith('HFGTYGLIDERTYPE:'):
            aircraft_type = line.split(':', 1)[1].strip()
        elif line.startswith('HFGIDGLIDERID:'):
            aircraft_reg = line.split(':', 1)[1].strip()
        elif line.startswith('B') and len(line) >= 35:
            b_records.append(line)

    if not flight_date or not b_records:
        return []

    def b_to_time(b: str) -> time:
        hh, mm, ss = int(b[1:3]), int(b[3:5]), int(b[5:7])
        return time(hh, mm, ss)

    takeoff_time = b_to_time(b_records[0])
    landing_time = b_to_time(b_records[-1])
    dur_s = (
        datetime.combine(flight_date, landing_time) - datetime.combine(flight_date, takeoff_time)
    ).seconds
    flight_time_min = dur_s // 60

    return [{
        'date': flight_date.isoformat(),
        'aircraft_type': aircraft_type,
        'aircraft_reg': aircraft_reg,
        'takeoff_time': takeoff_time.strftime('%H:%M'),
        'landing_time': landing_time.strftime('%H:%M'),
        'flight_time_min': flight_time_min,
        'source': 'igc',
    }]


def parse_file(source_type: str, content: bytes) -> list[dict]:
    """Parse uploaded file and return list of raw flight dicts."""
    if source_type == 'igc':
        return _parse_igc(content)
    if source_type in ('csv', 'leonardo', 'seeyou', 'weglide'):
        return _parse_csv_generic(content)
    if source_type == 'excel':
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            headers = [str(c.value or '').strip() for c in next(ws.iter_rows())]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, (str(v) if v is not None else '' for v in row))))
            return rows
        except ImportError:
            raise ImportServiceError('openpyxl is required for Excel imports.')
    raise ImportServiceError(f'Unsupported source type: {source_type}')


# ── Import flow ───────────────────────────────────────────────────────────────

def preview(db: Session, user: User, source_type: str, content: bytes) -> dict:
    """Parse file and categorise each row as new / duplicate / error."""
    raw_flights = parse_file(source_type, content)
    new_rows: list[dict] = []
    dup_rows: list[dict] = []
    error_rows: list[dict] = []

    for raw in raw_flights:
        try:
            flight_date = raw.get('date') or raw.get('Date') or raw.get('datum') or ''
            aircraft_reg = raw.get('aircraft_reg') or raw.get('Aircraft') or raw.get('registration') or ''
            takeoff_time = raw.get('takeoff_time') or raw.get('TakeoffTime') or raw.get('launch') or ''

            import_hash = _make_import_hash(str(user.id), flight_date, aircraft_reg, takeoff_time)
            exists = (
                db.query(Flight)
                .filter(
                    Flight.user_id == user.id,
                    Flight.raw_data['import_hash'].astext == import_hash,
                )
                .first()
            )
            if exists:
                dup_rows.append({**raw, '_import_hash': import_hash})
            else:
                new_rows.append({**raw, '_import_hash': import_hash})
        except Exception as exc:
            error_rows.append({**raw, '_error': str(exc)})

    return {
        'new': new_rows,
        'duplicate': dup_rows,
        'error': error_rows,
        'total': len(raw_flights),
    }


def confirm(
    db: Session,
    user: User,
    source_type: str,
    new_rows: list[dict],
    filename: Optional[str] = None,
) -> ImportLog:
    """Persist new_rows and create an ImportLog record. Caller commits."""
    log = ImportLog(
        user_id=user.id,
        source_type=source_type,
        filename=filename,
        imported_at=datetime.now(timezone.utc),
    )
    db.add(log)
    db.flush()

    new_count = 0
    dup_count = 0
    err_count = 0

    for raw in new_rows:
        try:
            import_hash = raw.get('_import_hash') or _make_import_hash(
                str(user.id),
                raw.get('date', ''),
                raw.get('aircraft_reg', ''),
                raw.get('takeoff_time', ''),
            )
            raw_data = {**raw, 'import_hash': import_hash}
            flight = Flight(
                id=uuid.uuid4(),
                user_id=user.id,
                source='import',
                import_id=log.id,
                raw_data=raw_data,
                synced_at=datetime.now(timezone.utc),
                date=_parse_date(raw) or date.today(),
                aircraft_type=raw.get('aircraft_type') or raw.get('Aircraft') or raw.get('glider'),
                aircraft_reg=raw.get('aircraft_reg') or raw.get('registration'),
                takeoff_time=_parse_time_str(raw.get('takeoff_time') or raw.get('TakeoffTime')),
                landing_time=_parse_time_str(raw.get('landing_time') or raw.get('LandingTime')),
                flight_time_min=_parse_int(raw.get('flight_time_min') or raw.get('Duration')),
                launch_type=raw.get('launch_type') or raw.get('LaunchType'),
                pilot=raw.get('pilot') or raw.get('Pilot'),
            )
            db.add(flight)
            new_count += 1
        except Exception as exc:
            logger.warning('Failed to import row: %s — %s', raw, exc)
            err_count += 1

    log.flights_new = new_count
    log.flights_dup = dup_count
    log.flights_error = err_count
    log.status = 'success' if err_count == 0 else 'partial'
    return log


def _make_import_hash(user_id: str, flight_date: str, aircraft_reg: str, takeoff_time: str) -> str:
    key = f'{user_id}:{flight_date}:{aircraft_reg}:{takeoff_time}'
    return hashlib.sha256(key.encode()).hexdigest()


def _parse_date(raw: dict) -> Optional[date]:
    for key in ('date', 'Date', 'datum', 'date_flight'):
        val = raw.get(key)
        if val:
            for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%m/%d/%Y'):
                try:
                    return datetime.strptime(str(val).strip(), fmt).date()
                except ValueError:
                    continue
    return None


def _parse_time_str(val) -> Optional[time]:
    if not val:
        return None
    try:
        parts = str(val).strip().split(':')
        return time(int(parts[0]), int(parts[1]))
    except Exception:
        return None


def _parse_int(val) -> Optional[int]:
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None
